import discord
import asyncio
import datetime
from openpyxl import load_workbook, Workbook

#--------------------------------Settings-----------------------------------------

intents = discord.Intents.default()
intents.members = True

now = datetime.datetime.now()
dateTime = now.strftime('%Y-%m-%d / ' + '%H:%M:%S')

wb = load_workbook("userDB.xlsx")
ws = wb.active
c_time = 1
c_name = 2
c_id = 3
c_status = 4

client = discord.Client(intents=intents)

#---------------------------------------------------------------------------------

#------------------------------Save Chat Log--------------------------------------

@client.event
async def on_message(message):
    if message.author.bot:
        return None
    channel = message.channel
    chat_log = open('chatLog.txt', 'a')
    chat_log.write("\n")
    chat_log.write(dateTime)
    chat_log.write("  Message from ( ")
    chat_log.write(channel.name)
    chat_log.write(" ) ")
    chat_log.write("{0.author}".format(message))
    chat_log.write(" : ")
    chat_log.write(message.content)
    chat_log.close()
    if (message.content == "!DB초기화"):
        delete()

#---------------------------------------------------------------------------------

#---------------------------Save User Info Log------------------------------------

@client.event   # 서버 신규 접속
async def on_member_join(member):
    signup(member.name, member.id)

@client.event   # 서버 퇴장
async def on_member_remove(member):
    withdraw(member.name, member.id)


def signup(_name, _id): # 입장
    _row = FinalRow()
    ws.cell(row=_row, column=c_time, value=dateTime)
    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value=("%d"%(_id)))
    ws.cell(row=_row, column=c_status, value="JOIN_IN")
    wb.save("userDB.xlsx")

def withdraw(_name, _id):   # 퇴장
    _row = FinalRow()
    ws.cell(row=_row, column=c_time, value=dateTime)
    ws.cell(row=_row, column=c_name, value=_name)
    ws.cell(row=_row, column=c_id, value=("%d"%(_id)))
    ws.cell(row=_row, column=c_status, value="LEFT_OUT")
    wb.save("userDB.xlsx")

def FinalRow():
    return ws.max_row + 1

def delete():
    ws.delete_rows(2,ws.max_row)
    wb.save("userDB.xlsx")

#---------------------------------------------------------------------------------

#---------------------------------------------------------------------------------



#로그랑 DB 볼 수 있는 GUI


client.run("ODQ0MjA5ODkxMzc1ODQxMjgw.YKPFug.TlpoWCLJbKvyIL9AYJCL008SrDE")


# 844202162908954627 채널 ID
# 844202162908954624 서버 ID
